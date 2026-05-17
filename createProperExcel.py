import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Create data with EXACT column names
data = {
    'customer_name': [
        'John Doe',
        'Jane Smith',
        'Robert Johnson',
        'Sarah Williams',
        'Michael Brown',
        'Emily Davis',
        'David Wilson',
        'Jessica Moore',
        'Christopher Taylor',
        'Amanda Anderson'
    ],
    'order_id': [
        'ORD-001',
        'ORD-002',
        'ORD-003',
        'ORD-004',
        'ORD-005',
        'ORD-006',
        'ORD-007',
        'ORD-008',
        'ORD-009',
        'ORD-010'
    ],
    'complaint_text': [
        'Product delivered on 2026-04-01. Item arrived damaged with missing components. Requested refund on 2026-04-05.',
        'Order shipped on 2026-04-10. Received wrong item. Submitted complaint on 2026-04-15. Need replacement or refund.',
        'Delivered on 2026-03-01. Quality not as described. Customer complained on 2026-05-05.',
        'Product delivered on 2026-04-20. Works but doesn\'t match product description. Complaint filed on 2026-05-01.',
        'Item arrived on 2026-02-01. Defective after 2 weeks of use. Refund requested on 2026-06-05.',
        'Delivery date was 2026-04-25. Customer satisfied initially but found issue later. Complaint on 2026-04-30.',
        'Package delivered on 2026-04-05. Missing items from order. Complaint submitted on 2026-04-10.',
        'Delivered on 2026-04-15. Color not as shown in pictures. Complaint filed on 2026-04-20.',
        'Item received on 2026-02-28. Stopped working after one month. Refund request on 2026-04-30.',
        'Shipped on 2026-04-10 and delivered on 2026-04-12. Size was incorrect. Complaint on 2026-05-05.'
    ],
    'complaint_date': [
        '2026-04-05',
        '2026-04-15',
        '2026-05-05',
        '2026-05-01',
        '2026-06-05',
        '2026-04-30',
        '2026-04-10',
        '2026-04-20',
        '2026-04-30',
        '2026-05-05'
    ]
}

# Create DataFrame
df = pd.DataFrame(data)

# Save to Excel
output_path = 'data/customer_complaints.xlsx'
df.to_excel(output_path, index=False, sheet_name='Complaints')

# Apply formatting
wb = load_workbook(output_path)
ws = wb.active

# Header formatting
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=12)

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Adjust column widths
ws.column_dimensions['A'].width = 18
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 60
ws.column_dimensions['D'].width = 15

# Format data rows
for row in ws.iter_rows(min_row=2, max_row=len(df)+1):
    for cell in row:
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

# Center align date columns
for row in ws.iter_rows(min_row=2, max_row=len(df)+1, min_col=4, max_col=4):
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='top')

# Set row heights
ws.row_dimensions[1].height = 25
for i in range(2, len(df)+2):
    ws.row_dimensions[i].height = 40

wb.save(output_path)
print(f"✅ Excel file created successfully at {output_path}")
print(f"Total records: {len(df)}")
print(f"\nColumns: {list(df.columns)}")